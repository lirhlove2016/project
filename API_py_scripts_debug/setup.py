# _*_ coding:utf-8 _*_
import os

"""
实现自动安装库文件1.xlrd，2.xlwt

"""
path="E:\download\setup"


def setup_openpyxl():
    try:
        from openpyxl import load_workbook
    except:
        print('\nnow install "openxl"')
        #set "setup.py" file path with yours
        setup = os.path.join(path, 'dependence', 'openpyxl-2.5.8', 'setup.py')        
        cwd = os.getcwd()
        os.chdir(os.path.dirname(setup))
        os.system(' '.join(('python', setup, 'install')))
        os.chdir(cwd)
        from openpyxl import load_workbook
        
    return load_workbook

def setup_xlrd():
    try:
        from xlrd import open_workbook
    except:
        print('\nnow install "xlrd"')
        setup = os.path.join(path, 'dependence', 'xlrd-1.1.0', 'setup.py')        
        cwd = os.getcwd()
        os.chdir(os.path.dirname(setup))
        os.system(' '.join(('python', setup, 'install')))
        os.chdir(cwd)
        from xlrd import load_workbook
        
    return open_workbook

def setup_xlwt():
    try:
        from xlwt import Workbook        
    except:
        print('\nnow install "xlwt"')
        setup = os.path.join(path, 'dependence', 'xlwt-1.3.0', 'setup.py')        
        cwd = os.getcwd()
        os.chdir(os.path.dirname(setup))
        os.system(' '.join(('python', setup, 'install')))
        os.chdir(cwd)
        from xlwt import Workbook
        
    return Workbook

if __name__ == '__main__':
    xlrd = setup_xlrd()
    xlwt = setup_xlwt()
    
